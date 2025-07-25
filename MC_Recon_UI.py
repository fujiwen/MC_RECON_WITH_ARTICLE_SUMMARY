import sys
import os
import pandas as pd
import numpy as np
import re
import logging
import configparser
from copy import copy
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.worksheet.properties import PageSetupProperties
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QPushButton, QTextEdit, QProgressBar, QFrame,
                             QFileDialog, QMessageBox, QListWidget, QListWidgetItem)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QRect
from PyQt5.QtGui import QFont, QPalette, QColor, QIcon
from PyQt5.QtWidgets import QDesktopWidget

class DataProcessThread(QThread):
    progress_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, input_files):
        super().__init__()
        self.input_files = input_files
        self.column_config = self.load_column_config()
    
    def excel_column_to_number(self, column_letter):
        """将Excel列字母转换为数字索引（从0开始）"""
        if isinstance(column_letter, int):
            return column_letter  # 如果已经是数字，直接返回
        
        column_letter = str(column_letter).strip()
        
        # 如果包含注释符号#，只取#前面的部分
        if '#' in column_letter:
            column_letter = column_letter.split('#')[0].strip()
        
        column_letter = column_letter.upper()
        
        # 如果是纯数字字符串，转换为整数
        if column_letter.isdigit():
            return int(column_letter)
        
        # 转换字母为数字
        result = 0
        for char in column_letter:
            if 'A' <= char <= 'Z':
                result = result * 26 + (ord(char) - ord('A') + 1)
            else:
                raise ValueError(f"无效的列标识符: {column_letter}")
        
        return result - 1  # 转换为从0开始的索引
    
    def load_column_config(self):
        """从配置文件加载列号配置"""
        config = configparser.ConfigParser()
        config_path = get_config_path()
        
        # 默认列号配置（使用字母格式）
        default_config = {
            'receipt_column': 'A',
            'supplier_column': 'D',
            'date_column': 'X',
            'product_name_column': 'A',
            'quantity_column': 'I',
            'unit_column': 'J',
            'unit_price_column': 'N',
            'subtotal_column': 'Z',
            'tax_amount_column': 'AE',
            'total_amount_column': 'AI',
            'department_column': 'AL'
        }
        
        if os.path.exists(config_path):
            try:
                config.read(config_path, encoding='utf-8')
                if 'Columns' in config:
                    # 从配置文件读取列号，如果不存在则使用默认值
                    for key in default_config:
                        if key in config['Columns']:
                            default_config[key] = config.get('Columns', key)
                logging.info(f'已加载列配置: {default_config}')
            except Exception as e:
                logging.error(f"读取列配置错误: {e}，使用默认配置")
        
        # 将所有列配置转换为数字索引
        numeric_config = {}
        for key, value in default_config.items():
            try:
                numeric_config[key] = self.excel_column_to_number(value)
                logging.info(f'{key}: {value} -> {numeric_config[key]}')
            except Exception as e:
                logging.error(f"转换列配置错误 {key}={value}: {e}")
                # 使用备用默认值
                backup_defaults = {
                    'receipt_column': 0, 'supplier_column': 3, 'date_column': 23,
                    'product_name_column': 0, 'quantity_column': 8, 'unit_column': 9,
                    'unit_price_column': 13, 'subtotal_column': 25, 'tax_amount_column': 30,
                    'total_amount_column': 34, 'department_column': 37
                }
                numeric_config[key] = backup_defaults.get(key, 0)
        
        return numeric_config
    
    def get_column_name(self, column_index):
        """根据列索引生成Unnamed列名"""
        return f'Unnamed: {column_index}'

    def format_mixed_text(self, text):
        if pd.isna(text):
            return text
        text = str(text)
        chinese_pattern = re.compile('[\u4e00-\u9fff]')
        match = chinese_pattern.search(text)
        if match:
            english_part = text[:match.start()].strip()
            chinese_part = text[match.start():].strip()
            if english_part and chinese_part:
                return f'{english_part}\n{chinese_part}'
        return text

    def extract_chinese(self, text):
        """提取文本中的中文字符"""
        if pd.isna(text):
            return text
        text = str(text)
        chinese_pattern = re.compile('[\u4e00-\u9fff]+')
        chinese_matches = chinese_pattern.findall(text)
        if chinese_matches:
            return ''.join(chinese_matches)
        return text

    def run(self):
        try:
            # 创建日志目录
            if not os.path.exists('logs'):
                os.makedirs('logs')
            
            # 配置日志
            log_filename = os.path.join('logs', f'process_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(log_filename, encoding='utf-8'),
                    logging.StreamHandler()
                ]
            )
            
            all_final_data = []
            
            for input_file in self.input_files:
                self.progress_signal.emit(f'开始读取文件：{os.path.basename(input_file)}')
                logging.info(f'开始读取文件：{input_file}')
                
                # 读取原始文件
                df = pd.read_excel(input_file, skiprows=8)
                logging.info(f'文件读取完成，共{len(df)}行数据')
                self.progress_signal.emit(f'文件读取完成，共{len(df)}行数据')
                
                # 获取收货单号的行索引
                receipt_column_name = self.get_column_name(self.column_config['receipt_column'])
                receipt_rows = df[df[receipt_column_name].astype(str).str.match(r'^(RTS)?000\d+$', na=False)].index
                
                # 创建一个空的列表来存储所有明细数据
                all_details = []
                
                # 遍历每个收货单号之间的行
                total_receipts = len(receipt_rows)
                for i in range(total_receipts):
                    start_idx = receipt_rows[i]
                    end_idx = receipt_rows[i+1] if i < len(receipt_rows)-1 else len(df)
                    
                    receipt = df.loc[start_idx, self.get_column_name(self.column_config['receipt_column'])]
                    supplier = df.loc[start_idx, self.get_column_name(self.column_config['supplier_column'])]
                    date = df.loc[start_idx, self.get_column_name(self.column_config['date_column'])]
                    
                    # 清理供应商名称和日期中的发票信息
                    if pd.notna(supplier):
                        supplier = re.sub(r'[（(].*[)）]|（专票.*|（普票.*|\s+专票.*|\s+普票.*|\d+%$', '', str(supplier)).strip()
                    
                    if pd.notna(date):
                        try:
                            date = pd.to_datetime(date)
                            if pd.notna(date):
                                date = date.strftime('%Y-%m-%d')
                        except:
                            date = None
                    
                    # 获取明细行（跳过收货单号行）
                    details = df.loc[start_idx+1:end_idx-1].copy()
                    
                    # 只保留非空行且不包含Page和Delivery Date的行
                    product_name_column = self.get_column_name(self.column_config['product_name_column'])
                    details = details[details[product_name_column].notna()]
                    details = details[~details[product_name_column].astype(str).str.contains('Page|Delivery Date', na=False)]
                    
                    if not details.empty:
                        details['收货单号'] = receipt
                        details['供应商名称'] = self.extract_chinese(supplier)
                        details['收货日期'] = date
                        details['商品名称'] = details[self.get_column_name(self.column_config['product_name_column'])].apply(self.format_mixed_text)
                        details['实收数量'] = details[self.get_column_name(self.column_config['quantity_column'])]
                        details['基本单位'] = details[self.get_column_name(self.column_config['unit_column'])]
                        details['单价'] = details[self.get_column_name(self.column_config['unit_price_column'])]
                        details['小计金额'] = details[self.get_column_name(self.column_config['subtotal_column'])]
                        details['税额'] = details[self.get_column_name(self.column_config['tax_amount_column'])]
                        details['税率'] = details[self.get_column_name(self.column_config['tax_amount_column'])] / details[self.get_column_name(self.column_config['subtotal_column'])]
                        details['小计价税'] = details[self.get_column_name(self.column_config['total_amount_column'])]
                        details['部门'] = details[self.get_column_name(self.column_config['department_column'])].apply(self.format_mixed_text)
                        
                        all_details.append(details[['收货单号', '收货日期', '商品名称', '实收数量', '基本单位',
                                                   '单价', '小计金额', '税额', '税率', '小计价税', '部门', '供应商名称']])
                    
                    progress = f'处理进度：{i+1}/{total_receipts}'
                    self.progress_signal.emit(progress)
                    logging.info(progress)
                
                # 合并所有明细数据
                if all_details:
                    file_df = pd.concat(all_details, ignore_index=True)
                    all_final_data.append(file_df)
                    logging.info(f'文件处理完成，共整理{len(file_df)}条记录')
                    self.progress_signal.emit(f'文件处理完成，共整理{len(file_df)}条记录')
            
            # 合并所有文件的数据
            final_df = pd.concat(all_final_data, ignore_index=True)
            logging.info(f'所有文件处理完成，共整理{len(final_df)}条记录')
            self.progress_signal.emit(f'所有文件处理完成，共整理{len(final_df)}条记录')
            
            # 创建供应商对账明细表文件夹
            if not os.path.exists('供应商对账明细'):
                os.makedirs('供应商对账明细')
                logging.info('创建供应商对账明细文件夹')
            
            # 按供应商名称分组并生成对账明细表
            total_suppliers = len(final_df['供应商名称'].unique())
            current_supplier = 0
            
            for supplier_name, supplier_data in final_df.groupby('供应商名称'):
                if pd.notna(supplier_name) and supplier_name.strip():
                    current_supplier += 1
                    self.progress_signal.emit(f'正在生成供应商对账单 ({current_supplier}/{total_suppliers}): {supplier_name}')
                    
                    # 按收货日期和收货单号排序
                    supplier_data = supplier_data.sort_values(['收货日期', '收货单号'])
                    
                    # 获取年月信息
                    first_date = pd.to_datetime(supplier_data['收货日期'].iloc[0])
                    year_month = first_date.strftime('%Y%m')
                    
                    # 创建年月目录
                    year_month_dir = os.path.join('供应商对账明细', year_month)
                    if not os.path.exists(year_month_dir):
                        os.makedirs(year_month_dir)
                    
                    # 计算合计金额
                    total_amount = supplier_data['小计价税'].sum()
                    
                    # 创建一个包含合计行的新数据框
                    summary_row = pd.DataFrame([{
                        '收货单号': '合计',
                        '收货日期': '',
                        '商品名称': '',
                        '实收数量': '',
                        '基本单位': '',
                        '单价': '',
                        '小计金额': supplier_data['小计金额'].sum(),
                        '税额': supplier_data['税额'].sum(),
                        '税率': '',
                        '小计价税': total_amount,
                        '部门': '',
                        '供应商名称': ''
                    }])
                    
                    supplier_data_with_summary = pd.concat([supplier_data, summary_row], ignore_index=True)
                    
                    # 创建新的Excel工作簿
                    wb = Workbook()
                    ws = wb.active
                    
                    # 设置页面布局
                    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4
                    ws.page_setup.fitToPage = True
                    ws.page_setup.fitToHeight = 0
                    ws.page_setup.fitToWidth = 1
                    ws.print_options.horizontalCentered = True
                    ws.print_options.verticalCentered = False
                    # 设置页面缩放比例为80%
                    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                    ws.sheet_view.zoomScale = 80
                    
                    # 读取配置文件中的公司名称
                    config = configparser.ConfigParser()
                    config_path = get_config_path()
                    company_name = 'HOTEL NAME'  # 默认值
                    if os.path.exists(config_path):
                        try:
                            config.read(config_path, encoding='utf-8')
                            if 'General' in config and 'company_name' in config['General']:
                                company_name = config['General']['company_name']
                        except Exception as e:
                            logging.error(f"读取配置文件错误: {e}")
                    
                    # 设置页脚文本、字体和大小
                    ws.oddFooter.center.text = f'\n第 &P 页，共 &N 页\n{company_name}'
                    ws.oddFooter.center.size = 11
                    ws.oddFooter.center.font = '微软雅黑'

                    
                    # 设置页边距（单位：厘米）
                    ws.page_margins = PageMargins(left=0.31, right=0.31, top=0.31, bottom=0.39, header=0.31, footer=0.21)
                    
                    # 设置列宽
                    column_widths = {
                        '收货单号': 17,
                        '收货日期': 19,
                        '商品名称': 60,
                        '实收数量': 19,
                        '基本单位': 19,
                        '单价': 20,
                        '小计金额': 14,
                        '税额': 14,
                        '税率': 10,
                        '小计价税': 20,
                        '部门': 35,
                        '供应商名称': 36
                    }
                    
                    # 设置酒店名称标题
                    hotel_title_row = 1
                    ws.merge_cells(start_row=hotel_title_row, start_column=1, end_row=hotel_title_row, end_column=len(column_widths))
                    hotel_title_cell = ws.cell(row=hotel_title_row, column=1, value='对账明细表')
                    hotel_title_cell.font = Font(name='微软雅黑', size=16, color='000000')
                    hotel_title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.row_dimensions[hotel_title_row].height = 22
                    
                    # 设置空白行2
                    blank_row2 = 2
                    ws.merge_cells(start_row=blank_row2, start_column=1, end_row=blank_row2, end_column=len(column_widths))
                    # 在A2单元格添加供应商名称信息
                    supplier_info = f'供应商名称：{supplier_name}'
                    blank_cell2 = ws.cell(row=blank_row2, column=1, value=supplier_info)
                    blank_cell2.font = Font(name='微软雅黑', size=13, color='000000')
                    blank_cell2.alignment = Alignment(horizontal='left', vertical='center')
                    ws.row_dimensions[blank_row2].height = 18.75

                    # 设置空白行3 - 添加对账周期信息
                    blank_row3 = 3
                    ws.merge_cells(start_row=blank_row3, start_column=1, end_row=blank_row3, end_column=len(column_widths))
                    
                    # 获取年月信息并计算月份第一天和最后一天
                    first_date = pd.to_datetime(supplier_data['收货日期'].iloc[0])
                    first_day = first_date.replace(day=1).strftime('%Y-%m-%d')
                    # 计算月份最后一天
                    if first_date.month == 12:
                        last_day_date = first_date.replace(year=first_date.year+1, month=1, day=1) - pd.Timedelta(days=1)
                    else:
                        last_day_date = first_date.replace(month=first_date.month+1, day=1) - pd.Timedelta(days=1)
                    last_day = last_day_date.strftime('%Y-%m-%d')
                    
                    # 设置对账周期信息
                    billing_cycle = f'对帐周期：{first_day} 至 {last_day}'
                    blank_cell3 = ws.cell(row=blank_row3, column=1, value=billing_cycle)
                    blank_cell3.font = Font(name='微软雅黑', size=13, color='000000')
                    blank_cell3.alignment = Alignment(horizontal='left', vertical='center')
                    ws.row_dimensions[blank_row3].height = 18.75

                    # 设置空白行4 - 添加小计金额合计信息
                    blank_row4 = 4
                    ws.merge_cells(start_row=blank_row4, start_column=1, end_row=blank_row4, end_column=len(column_widths))
                    
                    # 获取小计金额合计数据
                    total_subtotal = supplier_data['小计金额'].sum()
                    
                    # 设置小计金额信息
                    subtotal_info = f'Net净额：{total_subtotal:,.2f}'
                    blank_cell4 = ws.cell(row=blank_row4, column=1, value=subtotal_info)
                    blank_cell4.font = Font(name='微软雅黑', size=13, color='000000')
                    blank_cell4.alignment = Alignment(horizontal='left', vertical='center')
                    ws.row_dimensions[blank_row4].height = 18.75

                    # 设置空白行5 - 添加税额合计信息
                    blank_row5 = 5
                    ws.merge_cells(start_row=blank_row5, start_column=1, end_row=blank_row5, end_column=len(column_widths))
                    
                    # 获取税额合计数据
                    total_tax = supplier_data['税额'].sum()
                    
                    # 设置税额信息
                    tax_info = f'Vat税额：{total_tax:,.2f}'
                    blank_cell5 = ws.cell(row=blank_row5, column=1, value=tax_info)
                    blank_cell5.font = Font(name='微软雅黑', size=13, color='000000')
                    blank_cell5.alignment = Alignment(horizontal='left', vertical='center')
                    ws.row_dimensions[blank_row5].height = 18.75

                    # 设置空白行6 - 添加小计价税合计信息
                    blank_row6 = 6
                    ws.merge_cells(start_row=blank_row6, start_column=1, end_row=blank_row6, end_column=len(column_widths))
                    
                    # 获取小计价税合计数据
                    total_amount = supplier_data['小计价税'].sum()
                    
                    # 设置小计价税信息
                    gross_info = f'Gross含税总额：{total_amount:,.2f}'
                    blank_cell6 = ws.cell(row=blank_row6, column=1, value=gross_info)
                    blank_cell6.font = Font(name='微软雅黑', size=13, color='000000')
                    blank_cell6.alignment = Alignment(horizontal='left', vertical='center')
                    ws.row_dimensions[blank_row6].height = 18.75
                    
                    # 设置表头样式
                    header_font = Font(name='微软雅黑', size=13, bold=False, color='000000')
                    cell_font = Font(name='微软雅黑', size=13)
                    
                    # 设置对齐方式
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    right_alignment = Alignment(horizontal='right', vertical='center', shrink_to_fit=False)
                    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # 设置边框样式
                    thin_border = Border(
                        left=Side(style='hair', color='D3D3D3'),
                        right=Side(style='hair', color='D3D3D3'),
                        top=Side(style='hair', color='D3D3D3'),
                        bottom=Side(style='hair', color='D3D3D3')
                    )
                    header_border = Border(
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    # 合计行边框样式（只有上边框）
                    summary_border = Border(
                        top=Side(style='thin', color='000000')
                    )

                    # 写入表头
                    headers = list(supplier_data.columns)
                    
                    # 更新表头字段名称
                    header_mapping = {
                        '商品名称': '商品名称 Article',
                        '实收数量': '实收数量 QTY',
                        '基本单位': '基本单位 Unit',
                        '单价': '单价 Unit Price',
                        '小计金额': '净额 Net',
                        '税额': '税额 VAT',
                        '小计价税': '含税总额 Gross',
                        '部门': '成本中心 CostCenter'
                    }
                    
                    header_row = 7
                    for col, header in enumerate(headers, 1):
                        # 使用映射更新表头名称
                        display_header = header_mapping.get(header, header)
                        cell = ws.cell(row=header_row, column=col, value=display_header)
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.border = header_border
                        ws.column_dimensions[get_column_letter(col)].width = column_widths[header]
                        
                        # 隐藏供应商名称列
                        if header == '供应商名称':
                            ws.column_dimensions[get_column_letter(col)].hidden = True
                    
                    # 设置表头行高
                    ws.row_dimensions[header_row].height = 18.75
                    
                    # 冻结前七行
                    ws.freeze_panes = 'A8'

                    # 写入数据
                    for row_idx, row in enumerate(supplier_data.values, header_row + 1):
                        # 设置行高为40以适应双行文本
                        ws.row_dimensions[row_idx].height = 40
                        
                        # 检查是否为负数金额行
                        has_negative = False
                        for col_idx, value in enumerate(row, 1):
                            # 使用原始字段名称进行判断，因为headers中存储的是原始字段名
                            if headers[col_idx-1] in ['小计金额', '税额', '小计价税'] and pd.notna(value) and float(value) < 0:
                                has_negative = True
                                break
                        
                        # 设置斑马线效果（偶数行）
                        if row_idx % 2 == 0 and not has_negative:
                            row_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                        else:
                            row_fill = None
                        
                        # 写入单元格数据
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.font = cell_font
                            cell.border = thin_border
                            
                            # 如果是负数金额行，整行设置黄色背景
                            if has_negative:
                                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                # 使用原始字段名称进行判断，因为headers中存储的是原始字段名
                                if headers[col_idx-1] in ['小计金额', '税额', '小计价税'] and pd.notna(value) and float(value) < 0:
                                    cell.font = Font(name='微软雅黑', size=11, color='FF0000')
                            elif row_fill:
                                cell.fill = row_fill
                            
                            # 设置数字列的对齐方式和格式
                            # 使用原始字段名称进行判断，因为headers中存储的是原始字段名
                            if headers[col_idx-1] in ['商品名称', '部门']:
                                cell.alignment = wrap_alignment
                            elif headers[col_idx-1] in ['实收数量', '单价', '小计金额', '税额', '小计价税']:
                                cell.alignment = right_alignment
                                if pd.notna(value) and str(value).strip():
                                    cell.number_format = '#,##0.00'
                            elif headers[col_idx-1] == '税率':
                                cell.alignment = right_alignment
                                if pd.notna(value) and str(value).strip():
                                    cell.number_format = '0%'
                            else:
                                cell.alignment = center_alignment
                    
                    # 写入合计行
                    row_idx = len(supplier_data) + header_row + 1
                    for col_idx, value in enumerate(summary_row.iloc[0], 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = Font(name='微软雅黑', size=11, bold=True)
                        cell.border = summary_border
                        
                        # 设置数字列的对齐方式和格式
                        if headers[col_idx-1] in ['小计金额', '税额', '小计价税']:
                            cell.alignment = right_alignment
                            if pd.notna(value) and str(value).strip():
                                cell.number_format = '#,##0.00'
                        else:
                            cell.alignment = center_alignment
                    
                    # 设置重复打印的行
                    ws.print_title_rows = '1:7'

                    # 创建商品数量统计工作表
                    article_summary_ws = wb.create_sheet(title="Article_Summary")
                    
                    # 按商品名称分组统计数量
                    article_stats = supplier_data.groupby('商品名称').agg({
                        '实收数量': 'sum',
                        '基本单位': 'first',
                        '单价': 'mean',
                        '小计金额': 'sum',
                        '税额': 'sum',
                        '税率': 'first',
                        '小计价税': 'sum'
                    }).reset_index()
                    
                    # 按总数量降序排列
                    article_stats = article_stats.sort_values('实收数量', ascending=False)
                    
                    # 设置商品统计表的页面布局
                    article_summary_ws.page_setup.orientation = article_summary_ws.ORIENTATION_PORTRAIT
                    article_summary_ws.page_setup.paperSize = article_summary_ws.PAPERSIZE_A4
                    article_summary_ws.page_setup.fitToPage = True
                    article_summary_ws.page_setup.fitToHeight = 0
                    article_summary_ws.page_setup.fitToWidth = 1
                    article_summary_ws.print_options.horizontalCentered = True
                    article_summary_ws.print_options.verticalCentered = False
                    # 设置页面缩放比例为80%
                    article_summary_ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                    article_summary_ws.sheet_view.zoomScale = 80
                    
                    # 设置页脚
                    article_summary_ws.oddFooter.center.text = f'\n第 &P 页，共 &N 页\n{company_name}'
                    article_summary_ws.oddFooter.center.size = 11
                    article_summary_ws.oddFooter.center.font = '微软雅黑'
                    
                    # 设置页边距
                    article_summary_ws.page_margins = PageMargins(left=0.31, right=0.31, top=0.31, bottom=0.39, header=0.31, footer=0.21)
                    
                    # 设置商品统计表的列宽
                    summary_column_widths = {
                        '商品名称 Article': 58,
                        '总数量': 12,
                        '基本单位 Unit': 16,
                        '平均单价': 14,
                        '净额 Net': 14,
                        '税额 VAT': 14,
                        '税率': 10,
                        '含税总额 Gross': 19
                    }
                    
                    # 设置商品统计表标题
                    summary_title_row = 1
                    article_summary_ws.merge_cells(start_row=summary_title_row, start_column=1, end_row=summary_title_row, end_column=len(summary_column_widths))
                    summary_title_cell = article_summary_ws.cell(row=summary_title_row, column=1, value='商品数量统计表 Article Quantity Summary')
                    summary_title_cell.font = Font(name='微软雅黑', size=16, color='000000')
                    summary_title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    article_summary_ws.row_dimensions[summary_title_row].height = 22
                    
                    # 设置空白行
                    summary_blank_row = 2
                    article_summary_ws.merge_cells(start_row=summary_blank_row, start_column=1, end_row=summary_blank_row, end_column=len(summary_column_widths))
                    blank_cell = article_summary_ws.cell(row=summary_blank_row, column=1, value='')
                    blank_cell.font = Font(name='微软雅黑', size=20, color='000000')
                    blank_cell.alignment = Alignment(horizontal='center', vertical='center')
                    article_summary_ws.row_dimensions[summary_blank_row].height = 10
                    
                    # 写入商品统计表头
                    summary_headers = ['商品名称 Article', '总数量', '基本单位 Unit', '平均单价', '净额 Net', '税额 VAT', '税率', '含税总额 Gross']
                    summary_header_row = 3
                    for col, header in enumerate(summary_headers, 1):
                        cell = article_summary_ws.cell(row=summary_header_row, column=col, value=header)
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.border = header_border
                        article_summary_ws.column_dimensions[get_column_letter(col)].width = summary_column_widths[header]
                    
                    # 冻结前三行
                    article_summary_ws.freeze_panes = 'A4'
                    
                    # 写入商品统计数据
                    for row_idx, (_, row) in enumerate(article_stats.iterrows(), summary_header_row + 1):
                        # 设置行高
                        article_summary_ws.row_dimensions[row_idx].height = 40
                        
                        # 设置斑马线效果
                        if row_idx % 2 == 0:
                            row_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                        else:
                            row_fill = None
                        
                        # 写入数据
                        values = [
                            row['商品名称'],
                            row['实收数量'],
                            row['基本单位'],
                            row['单价'],
                            row['小计金额'],
                            row['税额'],
                            row['税率'],
                            row['小计价税']
                        ]
                        
                        for col_idx, value in enumerate(values, 1):
                            cell = article_summary_ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.font = cell_font
                            cell.border = thin_border
                            
                            if row_fill:
                                cell.fill = row_fill
                            
                            # 设置对齐方式和格式
                            # 由于summary_headers已更新，需要使用新的字段名称进行判断
                            if summary_headers[col_idx-1] == '商品名称 Article':
                                cell.alignment = wrap_alignment
                            elif summary_headers[col_idx-1] in ['总数量', '平均单价', '净额 Net', '税额 VAT', '含税总额 Gross']:
                                cell.alignment = right_alignment
                                if pd.notna(value) and str(value).strip():
                                    cell.number_format = '#,##0.00'
                            elif summary_headers[col_idx-1] == '税率':
                                cell.alignment = right_alignment
                                if pd.notna(value) and str(value).strip():
                                    cell.number_format = '0%'
                            else:
                                cell.alignment = center_alignment
                    
                    # 添加商品统计合计行
                    summary_total_row = len(article_stats) + summary_header_row + 1
                    summary_totals = [
                        '合计',
                        article_stats['实收数量'].sum(),
                        '',
                        '',
                        article_stats['小计金额'].sum(),
                        article_stats['税额'].sum(),
                        '',
                        article_stats['小计价税'].sum()
                    ]
                    
                    for col_idx, value in enumerate(summary_totals, 1):
                        cell = article_summary_ws.cell(row=summary_total_row, column=col_idx, value=value)
                        cell.font = Font(name='微软雅黑', size=11, bold=True)
                        cell.border = summary_border
                        
                        # 设置数字列的对齐方式和格式
                        if summary_headers[col_idx-1] in ['总数量', '平均单价', '小计金额', '税额', '小计价税']:
                            cell.alignment = right_alignment
                            if pd.notna(value) and str(value).strip():
                                cell.number_format = '#,##0.00'
                        else:
                            cell.alignment = center_alignment
                    
                    # 设置重复打印的行
                    article_summary_ws.print_title_rows = '1:3'
                    
                    # 保存文件
                    output_file = os.path.join(year_month_dir, f'{supplier_name}_对账明细.xlsx')
                    wb.save(output_file)
                    logging.info(f'已生成供应商对账单：{output_file}')
                    
                    # 不再生成独立的商品数量统计表文件
                    logging.info(f'商品数量统计表已添加到对账明细表中')
            
            # 创建备份文件夹
            if not os.path.exists('bak'):
                os.makedirs('bak')
                logging.info('创建备份文件夹')
            
            # 获取当前时间作为备份文件名
            current_time = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            
            # 备份数据
            backup_file = os.path.join('bak', f'cleaned_receiving_journal_{current_time}.xlsx')
            final_df.to_excel(backup_file, index=False)
            logging.info(f'数据已备份至：{backup_file}')
            
            self.progress_signal.emit('处理完成！')
            self.finished_signal.emit(True, '')
            
        except Exception as e:
            error_msg = f'处理过程中出现错误：{str(e)}'
            logging.error(error_msg)
            self.progress_signal.emit(error_msg)
            self.finished_signal.emit(False, error_msg)

class QTextEditLogger(logging.Handler):
    def __init__(self, widget):
        super().__init__()
        self.widget = widget
        self.widget.setReadOnly(True)
        self.widget.setFont(QFont('Helvetica', 16))  # 使用Helvetica字体
        
        # 设置样式
        self.widget.setStyleSheet("""
            QTextEdit {
                background-color: #2b2b2b;
                color: #ffffff;
                border: 1px solid #3c3c3c;
                border-radius: 5px;
                padding: 5px;
            }
        """)
        
        # 创建定时器用于更新日志
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_log)
        self.update_timer.start(100)  # 每100ms更新一次
        self.pending_messages = []

    def emit(self, record):
        msg = self.format(record)
        self.pending_messages.append(msg)

    def update_log(self):
        if self.pending_messages:
            for msg in self.pending_messages:
                self.widget.append(msg)
            self.pending_messages.clear()
            # 滚动到底部
            self.widget.verticalScrollBar().setValue(
                self.widget.verticalScrollBar().maximum()
            )

# 程序版本信息
VERSION = '1.1.18'

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.selected_files = []
        self.version = VERSION
        
        # 确保配置文件存在
        config_path = ensure_config_file()
        logging.info(f"配置文件路径：{config_path}")
        
        self.initUI()
        
        # 记录应用程序启动日志
        logging.info(f"应用程序启动，版本：{self.version}")
        
    def centerOnScreen(self):
        """将窗口居中显示在屏幕上"""
        # 获取屏幕几何信息
        screen_geometry = QDesktopWidget().availableGeometry()
        # 计算窗口居中位置
        x = (screen_geometry.width() - self.width()) // 2
        y = (screen_geometry.height() - self.height()) // 2
        # 移动窗口到居中位置
        self.move(x, y)
    

    def initUI(self):
        self.setWindowTitle(f'MC对帐明细工具 v{self.version}')
        # 设置全局窗口图标
        icon = QIcon(':/icons/app_icon')
        QApplication.setWindowIcon(icon)
        
        # 设置窗口大小
        self.resize(1024, 768)
        
        # 窗口居中显示
        self.centerOnScreen()
        
        # 创建主窗口部件和布局
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # 创建主布局
        layout = QVBoxLayout()
        
        # 创建左右分栏布局
        split_layout = QHBoxLayout()
        
        # 左侧：文件选择
        file_frame = QFrame()
        file_frame.setFrameShape(QFrame.StyledPanel)
        file_frame.setFrameShadow(QFrame.Raised)
        file_frame.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border-radius: 10px;
                padding: 15px;
                margin: 10px;
            }
        """)
        
        file_layout = QVBoxLayout()
        
        # 文件选择标题和按钮区域
        header_layout = QHBoxLayout()
        self.file_label = QLabel('已选择的文件：')
        self.file_label.setProperty('title', 'true')
        self.select_button = QPushButton('添加文件')
        self.select_button.setStyleSheet("""
            QPushButton {
                background-color: #4a90e2;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #357abd;
            }
            QPushButton:pressed {
                background-color: #2a5f9e;
            }
        """)
        self.select_button.clicked.connect(self.selectFiles)
        
        # 添加清空选择按钮
        self.clear_button = QPushButton('清空选择')
        self.clear_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        self.clear_button.clicked.connect(self.clearFiles)
        
        header_layout.addWidget(self.file_label)
        header_layout.addStretch()
        header_layout.addWidget(self.clear_button)
        header_layout.addWidget(self.select_button)
        
        # 文件列表
        self.file_list = QListWidget()
        self.file_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #e0e0e0;
                border-radius: 5px;
                padding: 5px;
                background-color: #f8f9fa;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #e0e0e0;
            }
            QListWidget::item:selected {
                background-color: #e3f2fd;
                color: #1976d2;
            }
        """)
        
        file_layout.addLayout(header_layout)
        file_layout.addWidget(self.file_list)
        file_frame.setLayout(file_layout)
        
        # 右侧：进度显示
        progress_frame = QFrame()
        progress_frame.setFrameShape(QFrame.StyledPanel)
        progress_frame.setFrameShadow(QFrame.Raised)
        progress_frame.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border-radius: 10px;
                padding: 15px;
                margin: 10px;
            }
        """)
        
        progress_layout = QVBoxLayout()
        progress_label = QLabel('处理进度')
        progress_label.setProperty('title', 'true')
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: none;
                border-radius: 10px;
                text-align: center;
                background-color: #f0f0f0;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #4caf50;
                border-radius: 10px;
            }
        """)
        self.progress_bar.setTextVisible(False)
        
        self.process_button = QPushButton('开始处理')
        self.process_button.setStyleSheet("""
            QPushButton {
                background-color: #4caf50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                margin-top: 10px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #43a047;
            }
            QPushButton:pressed {
                background-color: #388e3c;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.process_button.clicked.connect(self.startProcess)
        self.process_button.setEnabled(False)
        
        progress_layout.addWidget(progress_label)
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.process_button)
        progress_layout.addStretch()
        progress_frame.setLayout(progress_layout)
        
        # 设置左右分栏的比例（5:5）
        split_layout.addWidget(file_frame, 5)
        split_layout.addWidget(progress_frame, 5)
        
        # 下方：日志显示
        log_frame = QFrame()
        log_frame.setFrameShape(QFrame.StyledPanel)
        log_frame.setFrameShadow(QFrame.Raised)
        log_frame.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border-radius: 10px;
                padding: 15px;
                margin: 10px;
            }
        """)
        
        log_layout = QVBoxLayout()
        log_label = QLabel('处理日志')
        log_label.setProperty('title', 'true')
        self.progress_text = QTextEdit()
        self.progress_text.setReadOnly(True)
        
        log_layout.addWidget(log_label)
        log_layout.addWidget(self.progress_text)
        log_frame.setLayout(log_layout)
        
        # 添加所有部件到主布局（调整顺序，将日志放到下方）
        layout.addLayout(split_layout)
        layout.addWidget(log_frame)
        
        # 添加版权信息
        copyright_label = QLabel(f'Powered By Cayman Fu 2025 Ver {self.version}')
        copyright_label.setAlignment(Qt.AlignCenter)
        copyright_label.setStyleSheet("color: #666666;")
        layout.addWidget(copyright_label)
        
        main_widget.setLayout(layout)
        
        # 设置整体样式
        self.setStyleSheet("""
            * {
                font-family: "微软雅黑";
                font-size: 16px;
            }
            QPushButton {
                font-size: 14px;
            }
            QMainWindow {
                background-color: #f0f2f5;
            }
            QLabel {
                font-size: 16px;
            }
            QLabel[title="true"] {
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton {
                font-size: 16px;
            }
            QListWidget, QListWidget::item {
                font-size: 16px;
            }
            QTextEdit {
                font-size: 16px;
            }
            QProgressBar {
                font-size: 16px;
            }
            QMessageBox {
                font-size: 16px;
            }
            QMessageBox QLabel {
                font-size: 16px;
            }
            QMessageBox QPushButton {
                font-size: 16px;
                min-width: 70px;
                padding: 6px 12px;
            }
            QFileDialog {
                font-size: 16px;
            }
            QFileDialog QLabel {
                font-size: 16px;
            }
            QFileDialog QPushButton {
                font-size: 16px;
            }
            QFileDialog QComboBox {
                font-size: 16px;
            }
            QFileDialog QListView {
                font-size: 16px;
            }
        """)
    
    def selectFiles(self):
        # 尝试从上次的位置打开文件对话框
        last_dir = getattr(self, 'last_directory', '')
        if not last_dir or not os.path.exists(last_dir):
            last_dir = ''
            
        files, _ = QFileDialog.getOpenFileNames(
            self,
            '选择文件',
            last_dir,
            'Excel Files (*.xls *.xlsx);;All Files (*)'
        )
        if files:
            # 记住最后打开的目录
            self.last_directory = os.path.dirname(files[0])
            logging.info(f'选择了{len(files)}个文件')
            
            # 避免重复添加相同的文件
            new_files = [f for f in files if f not in self.selected_files]
            if new_files:
                self.selected_files.extend(new_files)
                self.updateFileList()
                self.process_button.setEnabled(True)
            else:
                warning_box = QMessageBox(self)
                warning_box.setWindowTitle('警告')
                warning_box.setText('所选文件已存在！')
                warning_box.setIcon(QMessageBox.Warning)
                warning_box.exec_()
    
    def clearFiles(self):
        """清空文件列表并重置界面状态"""
        self.selected_files.clear()
        self.updateFileList()
        self.process_button.setEnabled(False)
        logging.info('已清空文件列表')
    
    def updateFileList(self):
        self.file_list.clear()
        for file_path in self.selected_files:
            self.file_list.addItem(QListWidgetItem(file_path))
    
    def startProcess(self):
        if not self.selected_files:
            warning_box = QMessageBox(self)
            warning_box.setWindowTitle('警告')
            warning_box.setText('请先选择要处理的文件！')
            warning_box.setIcon(QMessageBox.Warning)
            warning_box.exec_()
            return
        
        self.process_button.setEnabled(False)
        self.select_button.setEnabled(False)
        self.clear_button.setEnabled(False)
        self.progress_text.clear()
        self.progress_bar.setRange(0, 0)  # 设置进度条为忙碌状态
        
        # 创建并启动处理线程
        self.process_thread = DataProcessThread(self.selected_files)
        self.process_thread.progress_signal.connect(self.updateProgress)
        self.process_thread.finished_signal.connect(self.processFinished)
        self.process_thread.start()
    
    def updateProgress(self, message):
        self.progress_text.append(message)
        # 滚动到底部
        self.progress_text.verticalScrollBar().setValue(
            self.progress_text.verticalScrollBar().maximum()
        )
    
    def processFinished(self, success, error_msg):
        self.progress_bar.setRange(0, 100)  # 恢复进度条正常状态
        self.progress_bar.setValue(100 if success else 0)
        self.process_button.setEnabled(True)
        self.select_button.setEnabled(True)
        self.clear_button.setEnabled(True)
        
        if success:
            # 获取处理的统计信息
            supplier_dir = '供应商对账明细'
            year_month_dirs = [d for d in os.listdir(supplier_dir) if os.path.isdir(os.path.join(supplier_dir, d))]
            
            if year_month_dirs:
                latest_dir = max(year_month_dirs)  # 获取最新的年月目录
                full_dir_path = os.path.join(supplier_dir, latest_dir)
                supplier_files = [f for f in os.listdir(full_dir_path) if f.endswith('.xlsx') and not f.startswith('~$')]
                
                stats_message = f'数据处理完成！\n\n处理结果:\n- 生成了{len(supplier_files)}个供应商对账单\n- 保存在目录: {full_dir_path}\n\n是否打开输出文件夹？'
            else:
                stats_message = '数据处理完成！是否打开输出文件夹？'
            
            info_box = QMessageBox(self)
            info_box.setWindowTitle('完成')
            info_box.setText(stats_message)
            info_box.setIcon(QMessageBox.Information)
            info_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            info_box.setDefaultButton(QMessageBox.Yes)
            reply = info_box.exec_()
            
            if reply == QMessageBox.Yes:
                # 使用跨平台的方法打开文件夹
                folder_path = os.path.abspath('供应商对账明细')
                try:
                    import subprocess
                    import webbrowser
                    
                    # 首先尝试使用平台特定的方法
                    if sys.platform == 'win32':
                        os.startfile(folder_path)  # Windows特有方法
                    elif sys.platform == 'darwin':  # macOS
                        subprocess.Popen(['open', folder_path])
                    elif sys.platform.startswith('linux'):  # Linux
                        subprocess.Popen(['xdg-open', folder_path])
                    else:
                        # 如果以上都不适用，尝试使用webbrowser模块
                        webbrowser.open('file://' + folder_path)
                except Exception as e:
                    logging.warning(f'无法打开文件夹：{e}')
                    warning_box = QMessageBox(self)
                    warning_box.setWindowTitle('提示')
                    warning_box.setText(f'无法自动打开文件夹，请手动查看：{folder_path}')
                    warning_box.setIcon(QMessageBox.Warning)
                    warning_box.exec_()
            # 处理完成后自动清空文件列表
            self.clearFiles()
            logging.info('处理完成，界面已重置')
        else:
            error_box = QMessageBox(self)
            error_box.setWindowTitle('错误')
            error_box.setText(f'处理失败：{error_msg}')
            error_box.setIcon(QMessageBox.Critical)
            error_box.exec_()
            logging.error(f'处理失败：{error_msg}')

def ensure_directories():
    """确保必要的目录结构存在"""
    required_dirs = ['logs', 'bak', '供应商对账明细']
    for directory in required_dirs:
        if not os.path.exists(directory):
            os.makedirs(directory)
            logging.info(f'创建目录: {directory}')

def get_app_dir():
    """获取应用程序所在目录，兼容打包后的exe和脚本运行模式"""
    if getattr(sys, 'frozen', False):
        # 如果是打包后的exe，使用sys.executable获取exe所在目录
        app_dir = os.path.dirname(sys.executable)
    else:
        # 如果是脚本运行，使用脚本所在目录
        app_dir = os.path.dirname(os.path.abspath(__file__))
    return app_dir

def get_config_path():
    """获取配置文件路径"""
    app_dir = get_app_dir()
    config_path = os.path.join(app_dir, 'config.ini')
    return config_path

def ensure_config_file():
    """确保配置文件存在，如果不存在则创建默认配置"""
    config_path = get_config_path()
    logging.info(f'配置文件路径: {config_path}')
    
    # 检查配置文件是否存在
    if not os.path.exists(config_path):
        logging.info('配置文件不存在，创建默认配置文件')
        config = configparser.ConfigParser()
        config['General'] = {
            'company_name': 'HOTEL NAME'
        }
        
        # 添加默认列号配置
        config['Columns'] = {
            'receipt_column': 'A',          # 收货单号 - Receipt Number Column
            'supplier_column': 'D',         # 供应商名称 - Supplier Column
            'date_column': 'X',             # 收货日期 - Date Column
            'product_name_column': 'A',     # 商品名称 Article - Product Name Column
            'quantity_column': 'I',         # 实收数量 Received Quantity Column
            'unit_column': 'J',             # 基本单位 Basic Unit Column
            'unit_price_column': 'N',       # 单价 Price Column
            'subtotal_column': 'Z',         # 小计金额 Subtotal Column
            'tax_amount_column': 'AE',      # 税额 Tax Column
            'total_amount_column': 'AI',    # 小计价税列 Total Amount Column
            'department_column': 'AL'       # 部门列 Department Column
        }
        
        # 写入配置文件
        try:
            with open(config_path, 'w', encoding='utf-8') as configfile:
                config.write(configfile)
            logging.info('默认配置文件创建成功，包含列号配置')
        except Exception as e:
            logging.error(f'创建配置文件失败: {e}')
    else:
        logging.info('配置文件已存在')
    
    return config_path

def check_expiration():
    """
    检查程序是否过期
    
    检查当前日期是否超过2025年12月31日。
    如果超过，则返回False，表示程序已过期；否则返回True。
    
    Returns:
        bool: 程序是否有效（True表示有效，False表示已过期）
    """
    # 获取当前日期
    current_date = datetime.now()
    # 设置过期日期为2025年12月31日
    expiration_date = datetime(2025, 12, 31)
    
    # 如果当前日期超过了2025年12月31日，则程序已过期
    if current_date > expiration_date:
        return False
        
    return True

def main():
    try:
        # 确保必要的目录存在
        ensure_directories()
        
        # 确保配置文件存在
        config_path = ensure_config_file()
        
        # 配置日志
        log_filename = os.path.join('logs', f'app_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        app = QApplication(sys.argv)
        # 导入资源文件并设置全局窗口图标
        import resources
        icon = QIcon(':/icons/app_icon')
        app.setWindowIcon(icon)
        
        # 确保任务栏图标与应用程序图标一致（Windows平台特定）
        if sys.platform == 'win32':
            import ctypes
            app_id = f'MC.ReconUI.{VERSION}'  # 应用程序ID
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
        
        # 检查程序是否过期
        if not check_expiration():
            logging.error('程序版本已过期，需要更新')
            error_box = QMessageBox(None)
            error_box.setWindowTitle('版本过期')
            error_box.setText('版本过低，请联系开发者Cayman更新')
            error_box.setIcon(QMessageBox.Critical)
            error_box.exec_()
            sys.exit(1)
        else:
            logging.info('程序版本检查通过')
        
        window = MainWindow()
        window.show()
        logging.info('应用程序启动成功')
        sys.exit(app.exec_())
    except Exception as e:
        logging.error(f'应用程序启动失败: {e}')
        error_box = QMessageBox(None)
        error_box.setWindowTitle('错误')
        error_box.setText(f'应用程序启动失败: {e}')
        error_box.setIcon(QMessageBox.Critical)
        error_box.exec_()
        sys.exit(1)

if __name__ == '__main__':
    main()